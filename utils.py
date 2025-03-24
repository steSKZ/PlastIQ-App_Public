import streamlit as st
from openpyxl import load_workbook
import pandas as pd

# Function to show dataframe
def show_dataframe(header: str, df: pd.DataFrame) -> None:
    st.write(header)
    st.dataframe(df)

def append_df_to_excel(file_path_input: str, df: pd.DataFrame, sheet_name: str, startrow: None | int = None, **to_excel_kwargs) -> None:
    try:
        # Try to open an existing workbook
        with pd.ExcelWriter(file_path_input, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # Get the last row in the existing Excel sheet if startrow is not defined
            if startrow is None:
                workbook = load_workbook(file_path_input)
                if sheet_name in workbook.sheetnames:
                    startrow = workbook[sheet_name].max_row
                else:
                    startrow = 0

            # Write out the new data below the existing data
            df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False, header=False, **to_excel_kwargs)

    except FileNotFoundError:
        # If the file does not exist, create it with the DataFrame
        with pd.ExcelWriter(file_path_input, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False, **to_excel_kwargs)
